import polars as pl
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def merge_multi_results(multi_results):
    """
    Merge multiple entity DataFrames from multi-mode analysis into a single DataFrame.
    Each result has a different groupby column, so we need to:
    1. Rename each groupby column to a common name
    2. Add a column indicating which groupby was used
    3. Convert all columns to string to avoid type conflicts
    4. Vertically stack all results
    """
    merged_dfs = []
    
    for result in multi_results:
        groupby_col = result["groupby_col"]
        
        # Extract anomalies data
        anomalies = result.get("anomalies", [])
        
        if not anomalies:
            continue
            
        # Convert anomalies to DataFrame
        df = pl.DataFrame(anomalies)
        
        # Rename the groupby column to a standard name
        if groupby_col in df.columns:
            df = df.rename({groupby_col: "ENTITY_ID"})
        
        # Add a column to track which grouping this came from
        df = df.with_columns(pl.lit(groupby_col).alias("ANALYSIS_GROUPBY"))
        
        # Convert all columns to string to avoid type conflicts when merging
        # This is necessary because different groupings may have different data types
        for col in df.columns:
            if col not in ["ANALYSIS_GROUPBY"]:  # Keep this one as-is
                df = df.with_columns(pl.col(col).cast(pl.Utf8).alias(col))
        
        # Ensure consistent column order
        # Move ANALYSIS_GROUPBY and ENTITY_ID to the front
        cols = df.columns
        front_cols = ["ANALYSIS_GROUPBY", "ENTITY_ID"]
        other_cols = [c for c in cols if c not in front_cols]
        df = df.select(front_cols + other_cols)
        
        merged_dfs.append(df)
    
    if not merged_dfs:
        # Return empty DataFrame with expected columns
        return pl.DataFrame({
            "ANALYSIS_GROUPBY": [],
            "ENTITY_ID": [],
            "ANOMALY_REASON": []
        })
    
    # Stack all DataFrames
    # Use diagonal=True to handle different columns across DataFrames
    # All columns are now strings, so no type conflicts
    result_df = pl.concat(merged_dfs, how="diagonal")
    
    # Fill nulls with empty string for better display
    result_df = result_df.fill_null("")
    
    return result_df


def create_aggregated_csv(entity_df):
    """Create CSV export of aggregated entity data"""
    csv_buffer = StringIO()
    entity_df.write_csv(csv_buffer)
    csv_buffer.seek(0)
    
    # Convert to BytesIO for consistency
    bytes_buffer = BytesIO(csv_buffer.getvalue().encode('utf-8'))
    bytes_buffer.seek(0)
    
    return bytes_buffer


def create_excel_report(raw_df, entity_df, anomaly_keys, id_col):
    """
    Create detailed Excel report with:
    1. Raw data sheet
    2. Aggregated data sheet
    3. Anomalies-only sheet
    """
    wb = Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    anomaly_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Sheet 1: Raw Data
    ws_raw = wb.create_sheet("Raw Data", 0)
    
    # Write headers
    for col_idx, column in enumerate(raw_df.columns, 1):
        cell = ws_raw.cell(row=1, column=col_idx, value=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_idx, row in enumerate(raw_df.iter_rows(), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws_raw.cell(row=row_idx, column=col_idx, value=value)
            if value is None or value == "":
                cell.value = ""
    
    # Auto-adjust column widths
    for col_idx in range(1, len(raw_df.columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in ws_raw[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_raw.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: Aggregated Data
    ws_agg = wb.create_sheet("Aggregated Data", 1)
    
    # Write headers
    for col_idx, column in enumerate(entity_df.columns, 1):
        cell = ws_agg.cell(row=1, column=col_idx, value=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data with highlighting
    for row_idx, row in enumerate(entity_df.iter_rows(), 2):
        # Check if this row is an anomaly
        is_anomaly = False
        entity_id_value = None
        
        for col_idx, (col_name, value) in enumerate(zip(entity_df.columns, row), 1):
            cell = ws_agg.cell(row=row_idx, column=col_idx, value=value)
            
            # Track entity ID
            if col_name == id_col or col_name == "ENTITY_ID":
                entity_id_value = value
            
            # Check if this is marked as anomaly
            if col_name == "STRONG_ANOMALY":
                # Handle both boolean and string representations
                if value in [True, "True", "true", 1, "1"]:
                    is_anomaly = True
            
            if value is None or value == "":
                cell.value = ""
        
        # Highlight anomaly rows
        if is_anomaly or (entity_id_value in anomaly_keys):
            for col_idx in range(1, len(entity_df.columns) + 1):
                ws_agg.cell(row=row_idx, column=col_idx).fill = anomaly_fill
    
    # Auto-adjust column widths
    for col_idx in range(1, len(entity_df.columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in ws_agg[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_agg.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 3: Anomalies Only
    # For merged results, all anomalies are already in entity_df
    # Just filter to show only the anomaly rows
    ws_anom = wb.create_sheet("Anomalies Only", 2)
    
    # Write headers
    for col_idx, column in enumerate(entity_df.columns, 1):
        cell = ws_anom.cell(row=1, column=col_idx, value=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write only anomaly data (entity_df already contains only anomalies in multi-mode)
    row_idx = 2
    for row in entity_df.iter_rows():
        for col_idx, value in enumerate(row, 1):
            cell = ws_anom.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = anomaly_fill
            if value is None or value == "":
                cell.value = ""
        row_idx += 1
    
    if row_idx == 2:
        # No data written, add a message
        ws_anom.cell(row=1, column=1, value="No anomalies detected")
    else:
        # Auto-adjust column widths
        for col_idx in range(1, len(entity_df.columns) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in ws_anom[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_anom.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer